"""
Sellvance - Motor de Relatorios Exportaveis
Gera relatorios em Excel (.xlsx), CSV e PDF com filtro de datas
"""
import io, csv, datetime
from database import get_db

# -- Excel (openpyxl) helpers --
def _excel_style_header(ws, headers, row=1):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    purple = PatternFill(start_color='6C63FF', end_color='6C63FF', fill_type='solid')
    white_bold = Font(bold=True, color='FFFFFF', size=11)
    thin = Side(style='thin', color='333333')
    border = Border(bottom=thin)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = white_bold
        cell.fill = purple
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    return row + 1

def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)

def _date_label(date_start, date_end):
    """Returns a human readable period label for report headers."""
    if date_start and date_end:
        return f'Periodo: {date_start} a {date_end}'
    elif date_start:
        return f'A partir de {date_start}'
    elif date_end:
        return f'Ate {date_end}'
    return 'Todos os dados'


# ══════════════════════════════════════════════════════════════════
# DASHBOARD REPORT
# ══════════════════════════════════════════════════════════════════
def generate_dashboard_report(org_id, fmt='xlsx', date_start='', date_end=''):
    db = get_db()

    # Build date filter for orders
    date_sql = ''
    params_base = [org_id]
    if date_start:
        date_sql += " AND date(ordered_at) >= date(?)"
        params_base.append(date_start)
    if date_end:
        date_sql += " AND date(ordered_at) <= date(?)"
        params_base.append(date_end)

    # KPIs
    rev = db.execute(f'SELECT COALESCE(SUM(revenue),0) as v FROM orders WHERE org_id=?{date_sql}', params_base).fetchone()['v']
    orders = db.execute(f'SELECT COUNT(*) as v FROM orders WHERE org_id=?{date_sql}', params_base).fetchone()['v']
    customers = db.execute(f'SELECT COUNT(DISTINCT contact_id) as v FROM orders WHERE org_id=?{date_sql}', params_base).fetchone()['v']
    avg_ticket = round(rev / max(orders, 1), 2)

    # Ads spend (use created_at for campaigns)
    ads_sql = ''
    ads_params = [org_id]
    if date_start:
        ads_sql += " AND date >= ?"
        ads_params.append(date_start)
    if date_end:
        ads_sql += " AND date <= ?"
        ads_params.append(date_end)
    spend = db.execute(f'SELECT COALESCE(SUM(spend),0) as v FROM ad_campaigns WHERE org_id=?{ads_sql}', ads_params).fetchone()['v']
    roas = round(rev / max(spend, 1), 2)

    # Revenue by channel
    channels = db.execute(
        f'SELECT channel, SUM(revenue) as rev, COUNT(*) as qty FROM orders WHERE org_id=?{date_sql} GROUP BY channel ORDER BY rev DESC',
        params_base
    ).fetchall()

    # Daily revenue (respect date filter or default 30 days)
    if date_start or date_end:
        daily = db.execute(
            f"SELECT date(ordered_at) as day, SUM(revenue) as rev FROM orders WHERE org_id=?{date_sql} GROUP BY day ORDER BY day",
            params_base
        ).fetchall()
    else:
        daily = db.execute(
            "SELECT date(ordered_at) as day, SUM(revenue) as rev FROM orders WHERE org_id=? AND ordered_at >= date('now','-30 days') GROUP BY day ORDER BY day",
            (org_id,)
        ).fetchall()

    period = _date_label(date_start, date_end)

    if fmt == 'csv':
        return _dashboard_csv(rev, orders, customers, avg_ticket, roas, spend, channels, daily, period)
    elif fmt == 'pdf':
        return _dashboard_pdf(rev, orders, customers, avg_ticket, roas, spend, channels, daily, period)
    else:
        return _dashboard_xlsx(rev, orders, customers, avg_ticket, roas, spend, channels, daily, period)


def _dashboard_xlsx(rev, orders, customers, avg_ticket, roas, spend, channels, daily, period=''):
    from openpyxl import Workbook
    from openpyxl.styles import Font
    wb = Workbook()

    ws = wb.active
    ws.title = 'KPIs'
    if period:
        ws.cell(row=1, column=1, value=period).font = Font(bold=True, size=12, color='6C63FF')
        start_row = 3
    else:
        start_row = 1
    _excel_style_header(ws, ['Metrica', 'Valor'], row=start_row)
    kpis = [
        ('Receita Total', f'R$ {rev:,.2f}'),
        ('Total de Pedidos', orders),
        ('Clientes Unicos', customers),
        ('Ticket Medio', f'R$ {avg_ticket:,.2f}'),
        ('Investimento Ads', f'R$ {spend:,.2f}'),
        ('ROAS', f'{roas}x'),
    ]
    for i, (k, v) in enumerate(kpis, start_row + 1):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v).font = Font(bold=True)
    _auto_width(ws)

    ws2 = wb.create_sheet('Receita por Canal')
    _excel_style_header(ws2, ['Canal', 'Receita', 'Pedidos'])
    for i, ch in enumerate(channels, 2):
        ws2.cell(row=i, column=1, value=ch['channel'] or 'Direto')
        ws2.cell(row=i, column=2, value=f"R$ {ch['rev']:,.2f}")
        ws2.cell(row=i, column=3, value=ch['qty'])
    _auto_width(ws2)

    ws3 = wb.create_sheet('Receita Diaria')
    _excel_style_header(ws3, ['Data', 'Receita'])
    for i, d in enumerate(daily, 2):
        ws3.cell(row=i, column=1, value=d['day'])
        ws3.cell(row=i, column=2, value=f"R$ {d['rev']:,.2f}")
    _auto_width(ws3)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, 'sellvance_dashboard.xlsx', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


def _dashboard_csv(rev, orders, customers, avg_ticket, roas, spend, channels, daily, period=''):
    buf = io.StringIO()
    w = csv.writer(buf)
    if period:
        w.writerow([period])
        w.writerow([])
    w.writerow(['=== DASHBOARD KPIs ==='])
    w.writerow(['Metrica', 'Valor'])
    w.writerow(['Receita Total', f'{rev:.2f}'])
    w.writerow(['Total Pedidos', orders])
    w.writerow(['Clientes Unicos', customers])
    w.writerow(['Ticket Medio', f'{avg_ticket:.2f}'])
    w.writerow(['Investimento Ads', f'{spend:.2f}'])
    w.writerow(['ROAS', f'{roas}x'])
    w.writerow([])
    w.writerow(['=== RECEITA POR CANAL ==='])
    w.writerow(['Canal', 'Receita', 'Pedidos'])
    for ch in channels:
        w.writerow([ch['channel'] or 'Direto', f"{ch['rev']:.2f}", ch['qty']])
    w.writerow([])
    w.writerow(['=== RECEITA DIARIA ==='])
    w.writerow(['Data', 'Receita'])
    for d in daily:
        w.writerow([d['day'], f"{d['rev']:.2f}"])

    out = io.BytesIO(buf.getvalue().encode('utf-8-sig'))
    out.seek(0)
    return out, 'sellvance_dashboard.csv', 'text/csv'


def _dashboard_pdf(rev, orders, customers, avg_ticket, roas, spend, channels, daily, period=''):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=40, bottomMargin=40)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title2', parent=styles['Title'], fontSize=20, textColor=colors.HexColor('#6C63FF'))
    subtitle = ParagraphStyle('Sub', parent=styles['Normal'], fontSize=12, textColor=colors.gray)

    elements = []
    elements.append(Paragraph('Sellvance - Relatorio Dashboard', title_style))
    elements.append(Paragraph(f'Gerado em {datetime.datetime.now().strftime("%d/%m/%Y %H:%M")}', subtitle))
    if period:
        elements.append(Paragraph(period, subtitle))
    elements.append(Spacer(1, 20))

    elements.append(Paragraph('KPIs Principais', styles['Heading2']))
    kpi_data = [
        ['Metrica', 'Valor'],
        ['Receita Total', f'R$ {rev:,.2f}'],
        ['Total Pedidos', str(orders)],
        ['Clientes Unicos', str(customers)],
        ['Ticket Medio', f'R$ {avg_ticket:,.2f}'],
        ['Investimento Ads', f'R$ {spend:,.2f}'],
        ['ROAS', f'{roas}x'],
    ]
    t = Table(kpi_data, colWidths=[200, 200])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6C63FF')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5FF')]),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 20))

    elements.append(Paragraph('Receita por Canal', styles['Heading2']))
    ch_data = [['Canal', 'Receita', 'Pedidos']]
    for ch in channels:
        ch_data.append([ch['channel'] or 'Direto', f"R$ {ch['rev']:,.2f}", str(ch['qty'])])
    t2 = Table(ch_data, colWidths=[150, 150, 100])
    t2.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6C63FF')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elements.append(t2)

    doc.build(elements)
    buf.seek(0)
    return buf, 'sellvance_dashboard.pdf', 'application/pdf'


# ══════════════════════════════════════════════════════════════════
# TRAFEGO PAGO REPORT
# ══════════════════════════════════════════════════════════════════
def generate_traffic_report(org_id, fmt='xlsx', date_start='', date_end=''):
    db = get_db()
    from traffic_ai import calc_metrics, score_campaign

    # Build date-filtered query
    sql = "SELECT * FROM ad_campaigns WHERE org_id=? AND platform IN ('meta','google')"
    params = [org_id]
    if date_start:
        sql += " AND date >= ?"
        params.append(date_start)
    if date_end:
        sql += " AND date <= ?"
        params.append(date_end)

    campaigns_raw = db.execute(sql, params).fetchall()

    campaigns = []
    platforms = {}
    for c in campaigns_raw:
        cd = dict(c)
        m = calc_metrics(cd)
        s = score_campaign(cd, m)
        action = 'Escalar' if s['score'] >= 75 else ('Otimizar' if s['score'] >= 50 else 'Pausar')
        row = {
            'name': cd.get('name',''), 'platform': cd.get('platform',''),
            'spend': cd.get('spend',0), 'revenue': cd.get('revenue',0),
            'impressions': cd.get('impressions',0), 'clicks': cd.get('clicks',0),
            'conversions': cd.get('conversions',0),
            'roas': m.get('roas',0), 'ctr': m.get('ctr',0), 'cpc': m.get('cpc',0),
            'cpa': m.get('cpa',0), 'score': s.get('score',0),
            'status': cd.get('status',''), 'action': action,
        }
        campaigns.append(row)
        p = cd.get('platform','outro')
        if p not in platforms:
            platforms[p] = {'spend': 0, 'revenue': 0, 'campaigns': 0}
        platforms[p]['spend'] += cd.get('spend', 0)
        platforms[p]['revenue'] += cd.get('revenue', 0)
        platforms[p]['campaigns'] += 1

    for p in platforms:
        platforms[p]['roas'] = round(platforms[p]['revenue'] / max(platforms[p]['spend'], 1), 2)

    period = _date_label(date_start, date_end)

    if fmt == 'csv':
        return _traffic_csv(campaigns, platforms, period)
    elif fmt == 'pdf':
        return _traffic_pdf(campaigns, platforms, period)
    else:
        return _traffic_xlsx(campaigns, platforms, period)


def _traffic_xlsx(campaigns, platforms, period=''):
    from openpyxl import Workbook
    from openpyxl.styles import Font
    wb = Workbook()

    ws = wb.active
    ws.title = 'Resumo Plataformas'
    if period:
        ws.cell(row=1, column=1, value=period).font = Font(bold=True, size=12, color='6C63FF')
        _excel_style_header(ws, ['Plataforma', 'Campanhas', 'Investimento', 'Receita', 'ROAS'], row=3)
        start = 4
    else:
        _excel_style_header(ws, ['Plataforma', 'Campanhas', 'Investimento', 'Receita', 'ROAS'])
        start = 2
    plat_names = {'meta': 'Meta Ads', 'google': 'Google Ads'}
    row = start
    for p, d in platforms.items():
        ws.cell(row=row, column=1, value=plat_names.get(p, p))
        ws.cell(row=row, column=2, value=d['campaigns'])
        ws.cell(row=row, column=3, value=f"R$ {d['spend']:,.2f}")
        ws.cell(row=row, column=4, value=f"R$ {d['revenue']:,.2f}")
        ws.cell(row=row, column=5, value=f"{d['roas']}x")
        row += 1
    _auto_width(ws)

    # Sheet 2: ALL campaigns (no limit)
    ws2 = wb.create_sheet('Todas as Campanhas')
    headers = ['Campanha', 'Plataforma', 'Status', 'Investimento', 'Receita', 'ROAS', 'CTR%', 'CPC', 'CPA', 'Impressoes', 'Cliques', 'Conversoes', 'Score IA', 'Acao']
    _excel_style_header(ws2, headers)
    for i, c in enumerate(sorted(campaigns, key=lambda x: x['score'], reverse=True), 2):
        ws2.cell(row=i, column=1, value=c['name'])
        ws2.cell(row=i, column=2, value=plat_names.get(c['platform'], c['platform']))
        ws2.cell(row=i, column=3, value=c['status'])
        ws2.cell(row=i, column=4, value=f"R$ {c['spend']:,.2f}")
        ws2.cell(row=i, column=5, value=f"R$ {c['revenue']:,.2f}")
        ws2.cell(row=i, column=6, value=f"{c['roas']}x")
        ws2.cell(row=i, column=7, value=f"{c['ctr']}%")
        ws2.cell(row=i, column=8, value=f"R$ {c['cpc']:,.2f}")
        ws2.cell(row=i, column=9, value=f"R$ {c['cpa']:,.2f}")
        ws2.cell(row=i, column=10, value=c['impressions'])
        ws2.cell(row=i, column=11, value=c['clicks'])
        ws2.cell(row=i, column=12, value=c['conversions'])
        ws2.cell(row=i, column=13, value=c['score'])
        ws2.cell(row=i, column=14, value=c['action'])
    _auto_width(ws2)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, 'sellvance_trafego_pago.xlsx', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


def _traffic_csv(campaigns, platforms, period=''):
    buf = io.StringIO()
    w = csv.writer(buf)
    if period:
        w.writerow([period])
        w.writerow([])
    w.writerow(['=== RESUMO POR PLATAFORMA ==='])
    w.writerow(['Plataforma', 'Campanhas', 'Investimento', 'Receita', 'ROAS'])
    plat_names = {'meta': 'Meta Ads', 'google': 'Google Ads'}
    for p, d in platforms.items():
        w.writerow([plat_names.get(p,p), d['campaigns'], f"{d['spend']:.2f}", f"{d['revenue']:.2f}", f"{d['roas']}x"])
    w.writerow([])
    w.writerow(['=== TODAS AS CAMPANHAS ==='])
    w.writerow(['Campanha', 'Plataforma', 'Status', 'Investimento', 'Receita', 'ROAS', 'CTR%', 'CPC', 'CPA', 'Impressoes', 'Cliques', 'Conversoes', 'Score', 'Acao'])
    for c in sorted(campaigns, key=lambda x: x['score'], reverse=True):
        w.writerow([c['name'], plat_names.get(c['platform'],c['platform']), c['status'],
                     f"{c['spend']:.2f}", f"{c['revenue']:.2f}",
                     f"{c['roas']}x", f"{c['ctr']}%", f"{c['cpc']:.2f}", f"{c['cpa']:.2f}",
                     c['impressions'], c['clicks'], c['conversions'], c['score'], c['action']])
    out = io.BytesIO(buf.getvalue().encode('utf-8-sig'))
    out.seek(0)
    return out, 'sellvance_trafego_pago.csv', 'text/csv'


def _traffic_pdf(campaigns, platforms, period=''):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=30, bottomMargin=30)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('T', parent=styles['Title'], fontSize=18, textColor=colors.HexColor('#6C63FF'))
    subtitle = ParagraphStyle('Sub', parent=styles['Normal'], fontSize=11, textColor=colors.gray)

    elements = []
    elements.append(Paragraph('Sellvance - Relatorio Trafego Pago', title_style))
    elements.append(Paragraph(f'Gerado em {datetime.datetime.now().strftime("%d/%m/%Y %H:%M")}', subtitle))
    if period:
        elements.append(Paragraph(period, subtitle))
    elements.append(Spacer(1, 15))

    plat_names = {'meta': 'Meta Ads', 'google': 'Google Ads'}
    plat_data = [['Plataforma', 'Campanhas', 'Investimento', 'Receita', 'ROAS']]
    for p, d in platforms.items():
        plat_data.append([plat_names.get(p,p), str(d['campaigns']), f"R$ {d['spend']:,.2f}", f"R$ {d['revenue']:,.2f}", f"{d['roas']}x"])
    t = Table(plat_data)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#6C63FF')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 15))

    # ALL campaigns (no 20 limit)
    elements.append(Paragraph('Todas as Campanhas', styles['Heading2']))
    camp_data = [['Campanha', 'Plat.', 'Invest.', 'Receita', 'ROAS', 'CTR', 'Score', 'Acao']]
    for c in sorted(campaigns, key=lambda x: x['score'], reverse=True):
        camp_data.append([c['name'][:30], plat_names.get(c['platform'],'')[:10], f"R${c['spend']:,.0f}",
                          f"R${c['revenue']:,.0f}", f"{c['roas']}x", f"{c['ctr']}%", str(c['score']), c['action']])
    t2 = Table(camp_data, colWidths=[170, 65, 75, 75, 45, 45, 40, 55])
    t2.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#6C63FF')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 8),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F5F5FF')]),
    ]))
    elements.append(t2)

    doc.build(elements)
    buf.seek(0)
    return buf, 'sellvance_trafego_pago.pdf', 'application/pdf'


# ══════════════════════════════════════════════════════════════════
# CRM / CONTATOS REPORT
# ══════════════════════════════════════════════════════════════════
def generate_crm_report(org_id, fmt='xlsx', date_start='', date_end=''):
    db = get_db()

    # Build date filter
    date_sql = ''
    params = [org_id]
    if date_start:
        date_sql += " AND date(last_order_at) >= date(?)"
        params.append(date_start)
    if date_end:
        date_sql += " AND date(last_order_at) <= date(?)"
        params.append(date_end)

    contacts = db.execute(
        f'SELECT name, email, phone, source, rfm_segment, ltv, total_orders, last_order_at FROM contacts WHERE org_id=?{date_sql} ORDER BY ltv DESC',
        params
    ).fetchall()

    segments = db.execute(
        f'SELECT rfm_segment, COUNT(*) as qty, SUM(ltv) as total_ltv FROM contacts WHERE org_id=?{date_sql} GROUP BY rfm_segment ORDER BY total_ltv DESC',
        params
    ).fetchall()

    total = len(contacts)
    total_ltv = sum(c['ltv'] or 0 for c in contacts)
    period = _date_label(date_start, date_end)

    if fmt == 'csv':
        return _crm_csv(contacts, segments, total, total_ltv, period)
    elif fmt == 'pdf':
        return _crm_pdf(contacts, segments, total, total_ltv, period)
    else:
        return _crm_xlsx(contacts, segments, total, total_ltv, period)


def _crm_xlsx(contacts, segments, total, total_ltv, period=''):
    from openpyxl import Workbook
    from openpyxl.styles import Font
    wb = Workbook()

    ws = wb.active
    ws.title = 'Contatos'
    if period:
        ws.cell(row=1, column=1, value=period).font = Font(bold=True, size=12, color='6C63FF')
        _excel_style_header(ws, ['Nome', 'Email', 'Telefone', 'Origem', 'Segmento RFM', 'LTV', 'Pedidos', 'Ultimo Pedido'], row=3)
        start = 4
    else:
        _excel_style_header(ws, ['Nome', 'Email', 'Telefone', 'Origem', 'Segmento RFM', 'LTV', 'Pedidos', 'Ultimo Pedido'])
        start = 2
    for i, c in enumerate(contacts, start):
        ws.cell(row=i, column=1, value=c['name'] or '')
        ws.cell(row=i, column=2, value=c['email'] or '')
        ws.cell(row=i, column=3, value=c['phone'] or '')
        ws.cell(row=i, column=4, value=c['source'] or '')
        ws.cell(row=i, column=5, value=c['rfm_segment'] or '')
        ws.cell(row=i, column=6, value=f"R$ {(c['ltv'] or 0):,.2f}")
        ws.cell(row=i, column=7, value=c['total_orders'] or 0)
        ws.cell(row=i, column=8, value=c['last_order_at'] or '')
    _auto_width(ws)

    ws2 = wb.create_sheet('Segmentacao RFM')
    _excel_style_header(ws2, ['Segmento', 'Quantidade', '% do Total', 'LTV Total', 'LTV Medio'])
    seg_names = {'champion': 'Campeoes', 'loyal': 'Leais', 'potential': 'Potenciais',
                 'new': 'Novos', 'at_risk': 'Em Risco', 'lost': 'Perdidos'}
    for i, s in enumerate(segments, 2):
        name = seg_names.get(s['rfm_segment'], s['rfm_segment'] or 'N/A')
        pct = round(s['qty'] / max(total, 1) * 100, 1)
        avg_ltv = round((s['total_ltv'] or 0) / max(s['qty'], 1), 2)
        ws2.cell(row=i, column=1, value=name)
        ws2.cell(row=i, column=2, value=s['qty'])
        ws2.cell(row=i, column=3, value=f"{pct}%")
        ws2.cell(row=i, column=4, value=f"R$ {(s['total_ltv'] or 0):,.2f}")
        ws2.cell(row=i, column=5, value=f"R$ {avg_ltv:,.2f}")
    _auto_width(ws2)

    ws3 = wb.create_sheet('Top 50 Clientes')
    _excel_style_header(ws3, ['#', 'Nome', 'Email', 'Segmento', 'LTV', 'Pedidos'])
    for i, c in enumerate(contacts[:50], 2):
        ws3.cell(row=i, column=1, value=i-1)
        ws3.cell(row=i, column=2, value=c['name'] or '')
        ws3.cell(row=i, column=3, value=c['email'] or '')
        ws3.cell(row=i, column=4, value=c['rfm_segment'] or '')
        ws3.cell(row=i, column=5, value=f"R$ {(c['ltv'] or 0):,.2f}")
        ws3.cell(row=i, column=6, value=c['total_orders'] or 0)
    _auto_width(ws3)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, 'sellvance_crm_contatos.xlsx', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


def _crm_csv(contacts, segments, total, total_ltv, period=''):
    buf = io.StringIO()
    w = csv.writer(buf)
    if period:
        w.writerow([period])
        w.writerow([])
    w.writerow(['=== CONTATOS ==='])
    w.writerow(['Nome', 'Email', 'Telefone', 'Origem', 'Segmento RFM', 'LTV', 'Pedidos', 'Ultimo Pedido'])
    for c in contacts:
        w.writerow([c['name'] or '', c['email'] or '', c['phone'] or '', c['source'] or '',
                     c['rfm_segment'] or '', f"{(c['ltv'] or 0):.2f}", c['total_orders'] or 0, c['last_order_at'] or ''])
    w.writerow([])
    w.writerow(['=== SEGMENTACAO RFM ==='])
    w.writerow(['Segmento', 'Quantidade', 'LTV Total'])
    for s in segments:
        w.writerow([s['rfm_segment'] or 'N/A', s['qty'], f"{(s['total_ltv'] or 0):.2f}"])

    out = io.BytesIO(buf.getvalue().encode('utf-8-sig'))
    out.seek(0)
    return out, 'sellvance_crm_contatos.csv', 'text/csv'


def _crm_pdf(contacts, segments, total, total_ltv, period=''):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=30, bottomMargin=30)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('T', parent=styles['Title'], fontSize=18, textColor=colors.HexColor('#6C63FF'))
    subtitle = ParagraphStyle('Sub', parent=styles['Normal'], fontSize=11, textColor=colors.gray)

    elements = []
    elements.append(Paragraph('Sellvance - Relatorio CRM', title_style))
    elements.append(Paragraph(f'Gerado em {datetime.datetime.now().strftime("%d/%m/%Y %H:%M")}', subtitle))
    if period:
        elements.append(Paragraph(period, subtitle))
    elements.append(Spacer(1, 10))

    elements.append(Paragraph(f'Total de contatos: {total} | LTV Total: R$ {total_ltv:,.2f}', styles['Normal']))
    elements.append(Spacer(1, 15))

    seg_names = {'champion': 'Campeoes', 'loyal': 'Leais', 'potential': 'Potenciais',
                 'new': 'Novos', 'at_risk': 'Em Risco', 'lost': 'Perdidos'}
    elements.append(Paragraph('Segmentacao RFM', styles['Heading2']))
    seg_data = [['Segmento', 'Qtd', '% Total', 'LTV Total']]
    for s in segments:
        name = seg_names.get(s['rfm_segment'], s['rfm_segment'] or 'N/A')
        pct = round(s['qty'] / max(total, 1) * 100, 1)
        seg_data.append([name, str(s['qty']), f'{pct}%', f"R$ {(s['total_ltv'] or 0):,.2f}"])
    t = Table(seg_data, colWidths=[120, 60, 70, 120])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#6C63FF')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 15))

    elements.append(Paragraph('Top 20 Clientes por LTV', styles['Heading2']))
    top_data = [['Nome', 'Segmento', 'LTV', 'Pedidos']]
    for c in contacts[:20]:
        top_data.append([
            (c['name'] or 'N/A')[:25],
            seg_names.get(c['rfm_segment'], c['rfm_segment'] or 'N/A'),
            f"R$ {(c['ltv'] or 0):,.2f}",
            str(c['total_orders'] or 0)
        ])
    t2 = Table(top_data, colWidths=[150, 90, 100, 60])
    t2.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#6C63FF')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F5F5FF')]),
    ]))
    elements.append(t2)

    doc.build(elements)
    buf.seek(0)
    return buf, 'sellvance_crm_contatos.pdf', 'application/pdf'
